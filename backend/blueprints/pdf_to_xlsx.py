import traceback
from io import BytesIO

import pdfplumber
from openpyxl import Workbook

from flask import Blueprint, request

from utils.helpers import error, send_file_and_cleanup

pdf_xlsx_bp = Blueprint("pdf_xlsx", __name__)

# openpyxl sheet titles are capped at 31 characters and can't contain
# any of []:*?/\\
_INVALID_SHEET_CHARS = ["[", "]", ":", "*", "?", "/", "\\"]


def _safe_sheet_title(title: str) -> str:
    for ch in _INVALID_SHEET_CHARS:
        title = title.replace(ch, "-")
    return title[:31] or "Sheet"


@pdf_xlsx_bp.route("/convertXlsx", methods=["POST"])
def convert_pdf_to_xlsx():
    try:
        if "file" not in request.files:
            return error("No file provided")

        pdf_file = request.files["file"]

        if pdf_file.filename == "":
            return error("No file selected")

        pdf_bytes = pdf_file.read()

        workbook = Workbook()
        workbook.remove(workbook.active)

        any_content = False

        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            if len(pdf.pages) == 0:
                return error("Empty PDF")

            for page_number, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()

                if tables:
                    for table_index, table in enumerate(tables, start=1):
                        if not table:
                            continue

                        sheet_title = _safe_sheet_title(
                            f"Page{page_number}"
                            if len(tables) == 1
                            else f"Page{page_number}_T{table_index}"
                        )
                        sheet = workbook.create_sheet(title=sheet_title)

                        for row in table:
                            sheet.append(
                                ["" if cell is None else str(cell) for cell in row]
                            )

                        _autosize_columns(sheet)
                        any_content = True
                else:
                    text = page.extract_text() or ""
                    if text.strip():
                        sheet_title = _safe_sheet_title(f"Page{page_number}")
                        sheet = workbook.create_sheet(title=sheet_title)
                        for line in text.splitlines():
                            sheet.append([line])
                        _autosize_columns(sheet)
                        any_content = True

        if not any_content:
            return error(
                "No tabular or text data could be extracted from this PDF. "
                "Scanned/image-only PDFs are not supported."
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        xlsx_bytes = output.getvalue()

        return send_file_and_cleanup(
            xlsx_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="converted.xlsx",
        )

    except Exception as e:
        traceback.print_exc()
        return error(str(e), 500)


def _autosize_columns(sheet, max_width=60):
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)

    for col, width in widths.items():
        sheet.column_dimensions[col].width = min(width + 2, max_width)